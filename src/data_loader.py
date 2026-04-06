import pandas as pd
import openpyxl

VALID_ADCS = [
    'Mylotarg', 'Adcetris', 'Kadcyla', 'Besponsa', 'Polivy', 'Padcev',
    'Enhertu', 'Trodelvy', 'Blenrep', 'Zynlonta', 'Tivdak', 'Elahere',
    'Datroway', 'Teliso-V'
]


def load_data(filepath: str) -> pd.DataFrame:
    """Load and clean the RÉSUMÉ_Régression sheet.

    Uses data_only=True to read cached formula values.
    Drops metadata rows at the bottom (non-ADC rows).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['RÉSUMÉ_Régression']

    headers = [cell.value for cell in ws[1]]
    rows = [
        list(row)
        for row in ws.iter_rows(min_row=2, values_only=True)
    ]
    df = pd.DataFrame(rows, columns=headers)

    # Keep only valid ADC rows
    df = df[df['ADC'].isin(VALID_ADCS)].reset_index(drop=True)

    # Cast numeric columns
    numeric_cols = ['P', 'D', 'H', 'B', 'L', 'E', 'V',
                    'S(payload,organe)', '%G≥3 observé',
                    'Y binaire (G≥3 >10%)', 'T-ADC v3 = Σ×V×S']
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    return df
